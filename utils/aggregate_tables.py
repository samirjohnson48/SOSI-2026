"""

"""

import pandas as pd
import numpy as np
from tqdm import tqdm
from openpyxl import load_workbook
from functools import reduce

from utils.stock_landings import get_numbers_from_string


def round_excel_file(filename, decimal_places=2, lt_one=False, skiprows=0):
    try:
        book = load_workbook(filename)
    except FileNotFoundError:
        print(f"Error: File '{filename}' not found.")
        return

    for sheet_name in book.sheetnames:
        sheet = book[sheet_name]

        for col in sheet.columns:
            for i, cell in enumerate(col):
                if i <= skiprows:
                    continue
                if cell.data_type == "n":
                    if isinstance(cell.value, int):
                        cell.number_format = "#,##0"
                    else:
                        if (
                            lt_one
                            and isinstance(cell.value, float)
                            and 0 < cell.value < 1
                        ):
                            cell.number_format = "< 1"
                        elif decimal_places == 0:
                            cell.number_format = "#,##0"
                        elif isinstance(cell.value, float) and 0 < cell.value < 1:
                            val_str = str(cell.value)
                            zeros = (
                                len(val_str.split(".")[1])
                                - len(val_str.split(".")[1].lstrip("0"))
                                if "." in val_str
                                else 0
                            )
                            cell.number_format = "#,##0." + "0" * (
                                zeros + decimal_places
                            )
                        else:
                            cell.number_format = "#,##0." + "0" * decimal_places

    book.save(filename)


def add_footnote(df, footnote_text, multi_index=False):
    if multi_index:
        levels = len(df.columns[0])
        footnote_df = pd.DataFrame(
            {tuple("" for _ in range(levels)): [footnote_text]}, index=["Footnote"]
        )
    else:
        footnote_df = pd.DataFrame({"": [footnote_text]}, index=["Footnote"])

    df_with_footnote = pd.concat([df, footnote_df])

    return df_with_footnote


def compute_count_for_group(df, group_col="Analysis Group", count_col="Tier"):
    counts = df.groupby(group_col)[count_col].value_counts().unstack(fill_value=0)

    total = counts.sum(numeric_only=True)

    counts.loc["Global"] = total

    rename_cols = {col: f"{count_col} {col}" for col in counts.columns}

    counts = counts.rename(columns=rename_cols)

    counts["Total"] = counts.sum(axis=1)

    return counts


def compute_status_by_number(data, group):
    grouped = (
        data.groupby(group)
        .agg(
            **{
                "No. of stocks": ("Status", "size"),
                "No. of U": ("Status", lambda x: (x == "U").sum()),
                "No. of MSF": ("Status", lambda x: (x == "M").sum()),
                "No. of O": ("Status", lambda x: (x == "O").sum()),
                "No. of Sustainable": (
                    "Status",
                    lambda x: ((x == "U") | (x == "M")).sum(),
                ),
                "No. of Unsustainable": ("Status", lambda x: (x == "O").sum()),
                "U (%)": ("Status", lambda x: (x == "U").mean() * 100),
                "MSF (%)": ("Status", lambda x: (x == "M").mean() * 100),
                "O (%)": ("Status", lambda x: (x == "O").mean() * 100),
                "Sustainable (%)": (
                    "Status",
                    lambda x: ((x == "M") | (x == "U")).mean() * 100,
                ),
                "Unsustainable (%)": ("Status", lambda x: (x == "O").mean() * 100),
            }
        )
        .reset_index()
    )

    # Set percentages to NaN when denominator (No. of stocks) is 0
    for col in ["U (%)", "MSF (%)", "O (%)", "Sustainable (%)", "Unsustainable (%)"]:
        grouped[col] = grouped.apply(
            lambda row: np.nan if row["No. of stocks"] == 0 else row[col], axis=1
        )

    # Add a final row with total values
    total_stocks = data.shape[0]
    totals = pd.DataFrame(
        {
            group: ["Global"],
            "No. of stocks": [total_stocks],
            "No. of U": [(data["Status"] == "U").sum()],
            "No. of MSF": [(data["Status"] == "M").sum()],
            "No. of O": [(data["Status"] == "O").sum()],
            "No. of Sustainable": [data["Status"].isin(["U", "M"]).sum()],
            "No. of Unsustainable": [(data["Status"] == "O").sum()],
            "U (%)": [
                (
                    np.nan
                    if total_stocks == 0
                    else ((data["Status"] == "U").sum() / total_stocks) * 100
                )
            ],
            "MSF (%)": [
                (
                    np.nan
                    if total_stocks == 0
                    else ((data["Status"] == "M").sum() / total_stocks) * 100
                )
            ],
            "O (%)": [
                (
                    np.nan
                    if total_stocks == 0
                    else ((data["Status"] == "O").sum() / total_stocks) * 100
                )
            ],
            "Sustainable (%)": [
                (
                    np.nan
                    if total_stocks == 0
                    else (data["Status"].isin(["U", "M"]).sum() / total_stocks) * 100
                )
            ],
            "Unsustainable (%)": [
                (
                    np.nan
                    if total_stocks == 0
                    else ((data["Status"] == "O").sum() / total_stocks) * 100
                )
            ],
        }
    )

    return pd.concat([grouped, totals], ignore_index=True)


def compute_weighted_status_by_number(data, tier_weights={1: 1, 2: 0.5, 3: 0.25}):
    sbn = pd.DataFrame()

    statuses = ["U", "M", "O"]

    for tier in [1, 2, 3]:
        tier_mask = data["Tier"] == tier
        temp = data[tier_mask].groupby("Analysis Group")["Status"].value_counts()
        temp = temp.reset_index().pivot(
            columns="Status", values="count", index="Analysis Group"
        )
        temp = temp.rename(columns={s: f"{s}_{tier}" for s in statuses})

        for s in statuses:
            temp[f"{s}_{tier}"] *= tier_weights[tier]

        if sbn.empty:
            sbn = temp
        else:
            sbn = pd.merge(sbn, temp, on="Analysis Group", how="outer")

    for s in statuses:
        sbn[s] = sbn[[f"{s}_{tier}" for tier in [1, 2, 3]]].sum(axis=1)

    sbn["T"] = sbn[statuses].sum(axis=1)

    total_row = sbn.sum().to_frame().T
    total_row.index = ["Global"]

    sbn = pd.concat([sbn, total_row])

    for s in statuses:
        sbn[f"{s} (%)"] = 100 * sbn[s] / sbn["T"]

    sbn["Sustainable (%)"] = sbn["U (%)"] + sbn["M (%)"]
    sbn["Unsustainable (%)"] = sbn["O (%)"]

    cols_to_keep = [f"{s} (%)" for s in statuses] + [
        "Sustainable (%)",
        "Unsustainable (%)",
    ]

    sbn = sbn[cols_to_keep]

    return sbn


def compare_status_by_number(update, previous):
    comparison = pd.merge(
        update,
        previous,
        on="Analysis Group",
        how="left",
        suffixes=("_update", "_previous"),
    )

    new_columns = []

    for col in comparison.columns:
        if col == "Analysis Group":
            new_columns.append(("", col))  # Keeping Area as a separate category
        elif col.endswith("_previous"):
            new_columns.append(
                ("Previous SoSI Categories", col.replace("_previous", ""))
            )
        elif col.endswith("_update"):
            new_columns.append(("Updated SoSI Categories", col.replace("_update", "")))

    comparison.columns = pd.MultiIndex.from_tuples(new_columns)

    return comparison


def compute_summary_of_stocks(data, group="Tier"):
    assessed_data_mask = data["Status"].isin(["U", "M", "O"])
    numeric_isscaap_mask = data["ISSCAAP Code"].apply(
        lambda x: isinstance(x, (int, float))
    )

    summary = (
        data.groupby(group)
        .agg(
            {
                group: "size",
                "Status": lambda x: x.isin(["U", "M", "O"]).sum(),
                "ASFIS Scientific Name": lambda x: x[assessed_data_mask].nunique(),
                "ISSCAAP Code": lambda x: x[
                    assessed_data_mask & numeric_isscaap_mask
                ].nunique(),
            }
        )
        .rename(
            columns={
                group: "Total stocks",
                "Status": "Total assessed stocks",
                "ASFIS Scientific Name": "Total ASFIS species (from total assessed stocks)",
                "ISSCAAP Code": "Total ISSCAAP Groups (from total assessed stocks)",
            }
        )
    )

    summary.loc["Global"] = summary.sum()
    summary.loc["Global", "Total ASFIS species (from total assessed stocks)"] = (
        data.loc[assessed_data_mask, "ASFIS Scientific Name"].nunique()
    )
    summary.loc["Global", "Total ISSCAAP Groups (from total assessed stocks)"] = (
        data.loc[assessed_data_mask & numeric_isscaap_mask, "ISSCAAP Code"].nunique()
    )

    return summary


def compute_summary_area_by_tier(data):
    tier_summaries = {}

    for tier in [1, 2, 3, "Total"]:
        tier_mask = (
            data["Tier"] == tier
            if tier != "Total"
            else pd.Series(True, index=data.index)
        )

        df = data[tier_mask].copy()

        sos = compute_summary_of_stocks(df, group="Analysis Group")

        rename_dict = {
            "Total assessed stocks": "No. of stocks",
            "Total ASFIS species (from total assessed stocks)": "No. of ASFIS species",
            "Total ISSCAAP Groups (from total assessed stocks)": "No. of ISSCAAP groups",
        }

        sos = sos.rename(columns=rename_dict).drop(columns=["Total stocks"])

        sbn = (
            compute_status_by_number(df, "Analysis Group")
            .set_index("Analysis Group")
            .drop(columns="No. of stocks")
        )

        comb = pd.merge(sos, sbn, left_index=True, right_index=True)

        key = f"Tier {tier}" if tier != "Total" else "Total"
        tier_summaries[key] = comb

    return tier_summaries


def compute_species_status_by_number(data, species_list, fishstat, year=2023):
    data = data[data["ASFIS Scientific Name"].isin(species_list)]
    group = (
        data.groupby(["ASFIS Scientific Name", "Status"]).size().unstack(fill_value=0)
    )
    global_totals = group.sum(axis=0)
    global_totals.name = "Global"
    group = pd.concat([group, global_totals.to_frame().T])
    total_counts = group.sum(axis=1)
    percentages = group.div(total_counts, axis=0) * 100
    landings = (
        fishstat[fishstat["ASFIS Scientific Name"].isin(species_list)]
        .groupby("ASFIS Scientific Name")[year]
        .sum()
        .sort_values(ascending=False)
    )
    landings = (
        landings.to_frame()
        .reset_index()
        .rename(columns={"ASFIS Scientific Name": "Species"})
        .set_index("Species")
    )
    result = pd.concat(
        [group, percentages, landings], axis=1, keys=["Counts", "%", "Landings"]
    )
    result.columns.names = ["Metric", "Status"]
    result = result.rename_axis("Species").reset_index()
    result.sort_values(("Landings", year), ascending=False, inplace=True)
    result.loc[result[("Species", "")] == "Global", ("Landings", year)] = result[
        ("Landings", year)
    ].sum()

    result[("Landings", year)] /= 1e6
    result = result.rename(columns={year: f"{year} (Mt)"}, level=1)

    return result


def create_sg_dict(
    sg_df,
    group_key="Analysis Group",
    area_key="FAO Area",
    species_key="ASFIS Scientific Name",
):
    sg_grouped = sg_df.groupby(group_key)

    sg_dict = {}

    for name, group in sg_grouped:
        sg_dict[name] = {}

        # Assign an Area to the special group if specified
        if all(group[area_key].notna()):
            if group[area_key].nunique() > 1:
                print(f"Cannot assign unique area to analysis group {name}")
                continue

            sg_dict[name][area_key] = group[area_key].loc[0]

        sg_dict[name][species_key] = list(group[species_key].unique())

    return sg_dict


def compute_total_area_landings(
    area,
    fishstat,
    species_landings,
    special_groups={},
    isscaap_to_remove=[],
    year_start=1950,
    year_end=2023,
    area_key="FAO Area",
    group_key="Analysis Group",
):
    sl = species_landings.copy()

    # Define special groups masks to either take out special group landings from FAO Areas
    # or calculate landings for special group categories

    sg_masks = {}

    for sg, sg_info in special_groups.items():
        sg_masks[sg] = {}

        sg_area = sg_info.get("FAO Area")
        sg_list = sg_info.get("ASFIS Scientific Name", [])

        sg_area_mask = (
            fishstat["Area"] == sg_area
            if sg_area
            else pd.Series(True, index=fishstat.index)
        )

        sg_mask_cap = fishstat["ASFIS Scientific Name"].isin(sg_list) & sg_area_mask

        sg_masks[sg]["fishstat"] = sg_mask_cap

        sg_mask_sl = sl["ASFIS Scientific Name"].isin(sg_list)

        if sg_area:
            sg_mask_sl = sg_mask_sl & sl[area_key] == sg_area

        sg_masks[sg]["sl"] = sg_mask_sl

    if area in special_groups.keys():
        cap = fishstat[sg_masks[area]["fishstat"]]

        years = list(range(year_start, year_end + 1))

        total_cap = cap[years].sum()

        return total_cap

    area_list = get_numbers_from_string(area) if isinstance(area, str) else [area]

    area_mask_cap = fishstat["Area"].isin(area_list)

    isscaap_mask_cap = ~fishstat["ISSCAAP Code"].isin(isscaap_to_remove)

    all_sg_mask_cap = pd.Series(False, index=area_mask_cap.index)

    for sg, masks_dict in sg_masks.items():
        all_sg_mask_cap = all_sg_mask_cap | masks_dict["fishstat"]

    cap = fishstat[area_mask_cap & isscaap_mask_cap & ~all_sg_mask_cap]

    # Add special group landings back to cap which appear reported in FAO Area in assessment

    sl_area_mask = sl[group_key] == area

    for sg in special_groups:
        sg_in_area = sl[sl_area_mask & sg_masks[sg]["sl"]]
        sg_in_area = sg_in_area.drop_duplicates(subset="ASFIS Scientific Name")

        if not sg_in_area.empty:
            cap = pd.concat([cap, sg_in_area])

    # Add landings from assessed stocks in ISSCAAP Groups which have been removed
    sl_isscaap_mask = sl["ISSCAAP Code"].isin(isscaap_to_remove)

    lta = sl[sl_area_mask & sl_isscaap_mask]

    if not lta.empty:
        cap = pd.concat([cap, lta])

    years = list(range(year_start, year_end + 1))

    total_cap = cap[years].sum()

    return total_cap


def compute_total_aquaculture_landings(
    area,
    aquaculture,
    special_groups={},
    isscaap_to_remove=[],
    year_start=1950,
    year_end=2023,
):
    sg_masks = {}

    for sg, sg_info in special_groups.items():
        sg_area = sg_info.get("FAO Area")
        sg_list = sg_info.get("ASFIS Scientific Name", [])

        sg_area_mask = (
            aquaculture["Area"] == sg_area
            if sg_area
            else pd.Series(True, index=aquaculture.index)
        )

        sg_mask_aqua = aquaculture["ASFIS Scientific Name"].isin(sg_list) & sg_area_mask

        sg_masks[sg] = sg_mask_aqua

    years = list(range(year_start, year_end + 1))

    if area in special_groups:
        aqua = aquaculture[sg_masks[area]]

        total_aqua = aqua[years].sum()

        return total_aqua

    area_list = get_numbers_from_string(area) if isinstance(area, str) else [area]

    area_mask = aquaculture["Area"].isin(area_list)

    isscaap_mask = ~aquaculture["ISSCAAP Code"].isin(isscaap_to_remove)

    all_sg_mask = pd.Series(False, index=area_mask.index)

    for sg, mask in sg_masks.items():
        all_sg_mask = all_sg_mask | mask

    aqua = aquaculture[area_mask & isscaap_mask & ~all_sg_mask]

    total_aqua = aqua[years].sum()

    return total_aqua


def compute_appendix_landings(
    species_landings,
    fishstat,
    aquaculture,
    isscaap_to_remove,
    isscaap_code_to_name,
    scientific_names,
    iso3_to_name,
    special_groups,
    year_start=1950,
    year_end=2023,
    last_decade_year=2020,
):
    sl_ = species_landings.copy()

    agg_dict = {
        "Status": "first",
        "Tier": "first",
        "ASFIS Name": "first",
        "ISSCAAP Code": "first",
        "FAO Area": list,
    }
    for year in range(year_start, year_end + 1):
        agg_dict[year] = "sum"

    sl = (
        sl_.groupby(["Analysis Group", "ASFIS Scientific Name", "Location"])
        .agg(agg_dict)
        .reset_index()
    )

    # Group the Status and Uncertainty by tier
    aggregated_status = sl.groupby(['Analysis Group', 'ASFIS Scientific Name'])[['Tier', 'Status']].value_counts().unstack(['Tier', 'Status'], fill_value=pd.NA)

    tiers = range(1,4)
    statuses = ['U', 'M', 'O']

    col_order = [(t,s) for t in tiers for s in statuses]

    new_multi_index = pd.MultiIndex.from_tuples(col_order, names=['Tier', 'Status'])

    aggregated_status = aggregated_status.reindex(columns=new_multi_index, fill_value=pd.NA)

    tier_cols = [(f"Tier {t}", s) for t in tiers for s in statuses]

    aggregated_status.columns = pd.MultiIndex.from_tuples(tier_cols)

    # Group the rest of the columns
    aggregated_species = (
        sl.groupby(["Analysis Group", "ASFIS Scientific Name"], dropna=False).agg(
            {
                "ASFIS Name": "first",
                "FAO Area": list,
                "ISSCAAP Code": "first",
                **{year: ["first", "sum"] for year in range(year_start, year_end + 1)},
            }
        )
    ).reset_index()

    aggregated_species[("FAO Area", "list")] = aggregated_species[("FAO Area", "list")].apply(
        lambda x: sum(x, [])
    )

    # Retrieve the most activate countries for each species for the given area(s)
    def most_active_countries(
        row, 
        country_key="ISO3", 
        year=year_end, 
        species_col=("ASFIS Scientific Name", ""),
        area_col=("FAO Area", "list")):
        species, area_list = row[species_col], row[area_col]

        if species not in scientific_names:
            return np.nan

        if ", " in species:
            species_list = species.split(", ")
            species_mask = fishstat["ASFIS Scientific Name"].isin(species_list)
        else:
            species_mask = fishstat["ASFIS Scientific Name"] == species

        area_mask = fishstat["Area"].isin(area_list)
        cap = fishstat[species_mask & area_mask][[country_key, year]]

        cap_countries = (
            cap.groupby(country_key)
            .sum()
            .sort_values(year, ascending=False)
            .reset_index()
        )
        cap_countries = cap_countries[cap_countries[year] > 0]
        cap_countries["Country"] = cap_countries[country_key].map(iso3_to_name)

        return ", ".join(cap_countries["Country"].values[:5])

    tqdm.pandas(desc="Retrieving Most Active Countries in 2023")

    aggregated_species[("Most Active Countries in 2023", "")] = aggregated_species[
        [("ASFIS Scientific Name", ""), ("FAO Area", "list")]
    ].progress_apply(most_active_countries, axis=1)
    
    # Merge the species and status data
    species_landings_dec = pd.merge(
        aggregated_species,
        aggregated_status,
        on=["Analysis Group", "ASFIS Scientific Name"],
    )

    new_cols = [col if isinstance(col[0], int) or "Tier" in col[0] else (col[0], "") for col in species_landings_dec.columns]

    species_landings_dec.columns = pd.MultiIndex.from_tuples(new_cols)

    sgs = [sg for sg in special_groups.keys() if "FAO Area" not in special_groups[sg]]

    for year in range(year_start, year_end + 1):
        # Total landings are sum for species in "Tuna", "Sharks" areas
        # since same species correspond to different areas
        species_landings_dec[("Production", year)] = species_landings_dec.apply(
            lambda row: (
                row[(year, "sum")]
                if row[("Analysis Group", "")] in sgs
                else row[(year, "first")]
            ),
            axis=1,
        )
        species_landings_dec.drop(
            columns=[(year, "first"), (year, "sum")], inplace=True
        )

    species_landings_dec = species_landings_dec.drop(columns=("FAO Area", ""))

    # Report in kilotonnes
    for year in range(year_start, year_end + 1):
        species_landings_dec[("Production", year)] /= 1e3

    # Create the decade columns for the appendix sheet
    def create_decade_cols(
        data,
        year_start=year_start,
        last_decade_year=last_decade_year,
    ):
        d = data.copy()
        for start in range(year_start, last_decade_year + 1, 10):
            end = start + 9

            if isinstance(data, pd.DataFrame):
                d[("Production (average per year)", f"{start}s")] = data.loc[:, [("Production", y) for y in range(start, end + 1)]].mean(axis=1)
            elif isinstance(data, pd.Series):
                d[("Production (average per year)", f"{start}s")] = data.loc[("Production", start):("Production", end)].mean()

        return d

    species_landings_dec = create_decade_cols(species_landings_dec)

    # Reorder columns
    columns_order = [
        ("Analysis Group", ""),
        ("ISSCAAP Code", ""),
        ("ASFIS Name", ""),
        ("ASFIS Scientific Name", ""),
        (f"Most Active Countries in {year_end}", ""),
    ]
    columns_order += sorted(
        [
            col
            for col in species_landings_dec.columns
            if isinstance(col[1], str) and col[1][:4].isdigit()
        ]
    )
    columns_order += [
        col for col in species_landings_dec.columns if isinstance(col[1], int)
    ]
    columns_order += tier_cols
    columns_order_mi = pd.MultiIndex.from_tuples(columns_order)

    species_landings_dec = species_landings_dec[columns_order_mi]

    # Retrieve numeric columns
    def get_numeric_cols(cols):
        return [
            col
            for col in cols
            if isinstance(col[1], (float, int))
            or (isinstance(col[1], str) and col[1][:4].isdigit())
        ]

    # Take out specificed ISSCAAP groups from aquaculture data
    aqua_isscaap_mask = ~aquaculture["ISSCAAP Code"].isin(isscaap_to_remove)
    aquaculture = aquaculture[aqua_isscaap_mask]

    # Build the appendix landings sheets
    # Data with decade columns
    summaries_w_dec = {}
    # Data with individual years
    summaries_w_year = {}

    for area in species_landings_dec[("Analysis Group", "")].unique():
        # Total assessed landings in area
        ag_mask = species_landings_dec[("Analysis Group", "")] == area
        area_landings = species_landings_dec[ag_mask].drop(columns=("Analysis Group", ""))

        # Create total rows for each ISSCAAP group
        isscaap_total = area_landings.groupby(("ISSCAAP Code", "")).sum().reset_index()
        isscaap_total[("ASFIS Name", "")] = isscaap_total[("ISSCAAP Code", "")].apply(
            lambda x: (
                str(int(x)) + f" - {isscaap_code_to_name.get(x, " ")}"
                if not pd.isna(x) and isinstance(x, (int, float))
                else x
            )
        )
        isscaap_total.loc[
            :,
            [
                ("ASFIS Scientific Name", ""),
                (f"Most Active Countries in {year_end}", ""),
            ]
            + tier_cols,
        ] = np.nan

        isscaap_grouped = (
            area_landings.groupby(("ISSCAAP Code", ""))[area_landings.columns]
            .apply(
                lambda group: pd.concat(
                    [
                        group,
                        isscaap_total[
                            isscaap_total[("ISSCAAP Code", "")]
                            == group[("ISSCAAP Code", "")].iloc[0]
                        ],
                    ],
                    ignore_index=True,
                ),
            )
            .reset_index(drop=True)
        )

        # Add missing ISSCAAP Group stocks to bottom
        no_isscaap_mask = area_landings[("ISSCAAP Code", "")].isna()

        if sum(no_isscaap_mask) > 0:
            area_landings[("ISSCAAP Code", "")] = area_landings[("ISSCAAP Code", "")].astype(object)
            area_landings.loc[no_isscaap_mask, ("ISSCAAP Code", "")] = "Missing"
            isscaap_grouped = pd.concat(
                [isscaap_grouped, area_landings[no_isscaap_mask]]
            )

        isscaap_grouped = isscaap_grouped[
            [col for col in isscaap_grouped.columns if col not in tier_cols] + tier_cols
        ]

        total_area = area_landings[get_numeric_cols(area_landings.columns)].sum()

        total_cap = (
            compute_total_area_landings(
                area,
                fishstat,
                species_landings,
                isscaap_to_remove=isscaap_to_remove,
                special_groups=special_groups,
            )
            / 1e3
        )
        
        total_cap.index = pd.MultiIndex.from_tuples([("Production", y) for y in total_cap.index])
        
        total_cap = create_decade_cols(total_cap)

        diff_cap = total_cap - total_area

        total_aqua = (
            compute_total_aquaculture_landings(
                area, aquaculture, species_landings, isscaap_to_remove=isscaap_to_remove
            )
            / 1e3
        )
        
        total_aqua.index = pd.MultiIndex.from_tuples([("Production", y) for y in total_aqua.index])

        total_aqua = create_decade_cols(total_aqua)

        total_production = total_cap + total_aqua

        total_area = total_area.to_frame().T
        total_area[("ASFIS Name", "")] = "Total selected species groups"
        total_cap = total_cap.to_frame().T
        total_cap[("ASFIS Name", "")] = "Total marine capture"
        diff_cap = diff_cap.to_frame().T
        diff_cap[("ASFIS Name", "")] = "Total other species groups"
        total_aqua = total_aqua.to_frame().T
        total_aqua[("ASFIS Name", "")] = "Total aquaculture"
        total_production = total_production.to_frame().T
        total_production[("ASFIS Name", "")] = "Total production"

        area_summary = pd.concat(
            [
                isscaap_grouped,
                total_area,
                diff_cap,
                total_cap,
                total_aqua,
                total_production,
            ]
        ).reset_index(drop=True)

        area_summary_dec = area_summary.drop(
            columns=[("Production", year) for year in list(range(year_start, last_decade_year + 10))]
        )

        dec_cols = [
            ("Production (average per year)", f"{start}s")
            for start in range(year_start, last_decade_year + 1, 10)
        ]
        area_summary_years = area_summary.drop(columns=dec_cols)

        summaries_w_dec[area] = area_summary_dec
        summaries_w_year[area] = area_summary_years

    return summaries_w_dec, summaries_w_year


def compute_weighted_percentages(
    stock_landings,
    key="Analysis Group",
    year=2023,
    landings_key="Stock Landings 2023",
):
    data = stock_landings.copy()

    # Group by key and Status to aggregate data
    group = data.groupby([key, "Status"])[landings_key].sum().unstack(fill_value=0)

    # Add a "Global" aggregation row
    global_totals = group.sum(axis=0)
    global_totals.name = "Global"
    group = pd.concat([group, global_totals.to_frame().T])

    # Calculate total landings per group
    total_landings = group.sum(axis=1).to_frame(name="Total Landings (Mt)")

    # Ensure required columns exist before computations
    for col in ["M", "U", "O"]:
        if col not in group.columns:
            group[col] = 0  # Add missing columns to avoid KeyErrors

    # Compute total sustainable and unsustainable landings
    total_landings["Sustainable (Mt)"] = (group["M"] + group["U"]) / 1e6
    total_landings["Unsustainable (Mt)"] = group["O"] / 1e6
    total_landings["MSF (Mt)"] = group["M"] / 1e6
    total_landings["U (Mt)"] = group["U"] / 1e6
    total_landings["O (Mt)"] = group["O"] / 1e6

    # Ensure no division by zero
    wp = group.div(group.sum(axis=1).replace(0, 1), axis=0) * 100

    # Compute weighted percentages
    wp["Sustainable (%)"] = wp["M"] + wp["U"]
    wp["Unsustainable (%)"] = wp["O"]
    wp.rename(columns={"U": "U (%)", "M": "MSF (%)", "O": "O (%)"}, inplace=True)

    # Organize and rename columns
    total_landings = total_landings[
        ["U (Mt)", "MSF (Mt)", "O (Mt)", "Sustainable (Mt)", "Unsustainable (Mt)"]
    ]
    wp = wp[["U (%)", "MSF (%)", "O (%)", "Sustainable (%)", "Unsustainable (%)"]]

    # Combine totals and percentages
    result = pd.concat(
        [total_landings, wp],
        axis=1,
        keys=["Total Landings", "Weighted % by Landings"],
    )

    result.index.name = key

    return result


def get_weighted_percentages_and_total_landings(
    weighted_percentages,
    fishstat,
    species_landings,
    special_groups={},
    isscaap_to_remove=[],
    year=2023,
):
    wp = weighted_percentages.copy()
    wp_cols = [col[1] for col in wp.columns]
    wp.columns = wp_cols

    wp["Total Assessed Landings (Mt)"] = (
        wp["Sustainable (Mt)"] + wp["Unsustainable (Mt)"]
    )
    pc_cols = [col for col in wp_cols if "%" in col]
    cols_to_keep = ["Total Assessed Landings (Mt)"] + pc_cols

    wp_totl = wp[cols_to_keep]

    wp_totl = wp_totl.reset_index()

    wp_totl["Total Landings (Mt)"] = wp_totl["Analysis Group"].apply(
        compute_total_area_landings,
        args=(
            fishstat,
            species_landings,
            special_groups,
            isscaap_to_remove,
            year,
            year,
        ),
    )

    wp_totl["Total Landings (Mt)"] /= 1e6

    wp_totl = wp_totl[
        ["Analysis Group", "Total Landings (Mt)", "Total Assessed Landings (Mt)"]
        + pc_cols
    ]

    wp_totl = wp_totl.set_index("Analysis Group")

    wp_totl.loc["Global", "Total Landings (Mt)"] = wp_totl["Total Landings (Mt)"].sum()

    new_cols = [
        ("Weighted % by Landings", col) if "%" in col else ("", col)
        for col in wp_totl.columns
    ]
    wp_totl.columns = pd.MultiIndex.from_tuples(new_cols)

    return wp_totl


def get_weighted_percentages_by_tier_and_area(stock_landings, total_landings):
    areas = stock_landings["Analysis Group"].unique()
    areas_df = pd.DataFrame()

    tl_cols = [("", "", "Total Landings (Mt)")]

    def wp_tier(stock_landings, area=None):
        if area:
            area_mask = stock_landings["Analysis Group"] == area
        else:
            area_mask = pd.Series(True, index=stock_landings.index)
            area = "Global"

        d = compute_weighted_percentages(stock_landings[area_mask], key="Tier")

        d2_cols = (
            [("", "", "Analysis Group")]
            + tl_cols
            + [
                (f"Tier {i}", col[0], col[1])
                for i in d.index
                if isinstance(i, int)
                for col in d.columns
            ]
        )
        d2 = pd.DataFrame(columns=pd.MultiIndex.from_tuples(d2_cols))

        d2.loc[0, ("", "", "Analysis Group")] = area

        d2.loc[0, tl_cols] = total_landings.loc[
            area, [(col[0], col[2]) for col in tl_cols]
        ].values

        d2 = d2.rename(
            columns={"Total Landings (Mt)": "Total Landings in Area (Mt)"}, level=2
        )

        for i in range(1, 4):
            cols = [col for col in d2.columns if col[0] == f"Tier {i}"]
            if i in d.index:
                d2.loc[0, cols] = d.loc[i].values
                d2[(f"Tier {i}", "", "Total Landings (Mt)")] = (
                    d2.loc[0, (f"Tier {i}", "Total Landings", "Sustainable (Mt)")]
                    + d2.loc[0, (f"Tier {i}", "Total Landings", "Unsustainable (Mt)")]
                )
            else:
                d2.loc[0, cols] = [0] * len(cols)
                d2[(f"Tier {i}", "", "Total Landings (Mt)")] = 0

        cols_to_drop = [
            "U (Mt)",
            "MSF (Mt)",
            "O (Mt)",
            "Sustainable (Mt)",
            "Unsustainable (Mt)",
        ]
        d2 = d2.drop(
            columns=[
                (f"Tier {i}", "Total Landings", col)
                for i in range(1, 4)
                if i in d.index
                for col in cols_to_drop
            ]
        )
        tier1_cols = [("Tier 1", "", "Total Landings (Mt)")] if 1 in d.index else []
        tier2_cols = [("Tier 2", "", "Total Landings (Mt)")] if 2 in d.index else []
        tier3_cols = [("Tier 3", "", "Total Landings (Mt)")] if 3 in d.index else []

        tier1_cols += [
            col
            for col in d2.columns
            if col[0] == "Tier 1" and col[1] == "Weighted % by Landings"
        ]
        tier2_cols += [
            col
            for col in d2.columns
            if col[0] == "Tier 2" and col[1] == "Weighted % by Landings"
        ]
        tier3_cols += [
            col
            for col in d2.columns
            if col[0] == "Tier 3" and col[1] == "Weighted % by Landings"
        ]
        col_sort = (
            [("", "", "Analysis Group"), ("", "", "Total Landings in Area (Mt)")]
            + tier1_cols
            + tier2_cols
            + tier3_cols
        )

        return d2[col_sort]

    for area in areas:
        area_df = wp_tier(stock_landings, area)
        areas_df = pd.concat([areas_df, area_df])

    global_df = wp_tier(stock_landings)

    areas_df = pd.concat([areas_df, global_df])

    return areas_df


def compute_percent_coverage(
    stock_landings,
    species_landings,
    fishstat,
    isscaap_to_remove,
    landings_key="Stock Landings 2023",
    tier=None,
    year=2023,
    area_key="FAO Area",
):
    total_cov, total_area_l = 0, 0
    pc_dict = {}

    areas = stock_landings[area_key].unique()

    for area in areas:
        tier_mask = (
            stock_landings["Tier"] == tier
            if tier
            else pd.Series(True, index=stock_landings.index)
        )

        area_mask = stock_landings[area_key] == area

        cov = stock_landings[tier_mask & area_mask][landings_key].sum()

        area_l = compute_total_area_landings(
            area,
            fishstat,
            species_landings,
            isscaap_to_remove=isscaap_to_remove,
            special_groups={},
            area_key=area_key,
        )[year]

        pc_dict[area] = 100 * cov / area_l

        total_cov += cov
        total_area_l += area_l

    pc_dict["Global"] = 100 * total_cov / total_area_l

    pc = pd.DataFrame(pc_dict, index=["Coverage (%)"]).T.reset_index(names="FAO Area")

    return pc


def compute_percent_coverage_tiers(
    stock_landings,
    species_landings,
    fishstat,
    isscaap_to_remove,
):
    pc_tiers = pd.DataFrame()

    for tier in [1, 2, 3]:
        pc_tier = compute_percent_coverage(
            stock_landings,
            species_landings,
            fishstat,
            isscaap_to_remove,
            tier=tier,
        )

        pc_tier = pc_tier.rename(columns={"Coverage (%)": f"Coverage (%) Tier {tier}"})

        if pc_tiers.empty:
            pc_tiers = pc_tier.copy()
        else:
            pc_tiers = pd.merge(pc_tiers, pc_tier, on="FAO Area")

    pc_tiers = pc_tiers.set_index("FAO Area")

    pc_tiers["Coverage (%) Total"] = pc_tiers.sum(axis=1)

    return pc_tiers.reset_index()


def compare_weighted_percentages(previous, update):
    comp = pd.merge(update, previous, left_index=True, right_index=True, how="left")

    comp = comp.drop(columns=[col for col in comp.columns if "(Mt)" in col[1]])

    comp.columns = pd.MultiIndex.from_tuples(
        [
            (
                ("Updated SoSI Categories", col[1])
                if "_x" in col[0]
                else ("Previous SoSI Categories", col[1])
            )
            for col in comp.columns
        ]
    )

    return comp


def compute_species_weighted_percentages(stock_landings, species_list):
    species_mask = stock_landings["ASFIS Scientific Name"].isin(species_list)
    species_data = stock_landings[species_mask]

    group = (
        species_data.groupby(["ASFIS Scientific Name", "Status"])["Stock Landings 2023"]
        .sum()
        .unstack(fill_value=0)
    )

    global_totals = group.sum(axis=0)
    global_totals.name = "Global"
    group = pd.concat([group, global_totals.to_frame().T])

    total_landings = group.sum(axis=1)
    weighted_percentages = group.div(total_landings, axis=0) * 100

    result = pd.concat(
        [group, weighted_percentages],
        axis=1,
        keys=["Total Landings", "Weighted % by Landings"],
    )

    result.columns.names = ["Metric", "Status"]
    result = result.rename_axis("Species").reset_index()

    for status in ["M", "O", "U"]:
        result[("Total Landings", status)] /= 1e6
    result = result.rename(columns={"Total Landings": "Total Landings (Mt)"}, level=0)

    return result


def compute_top_species_by_area(
    ags, stock_assessments, stock_landings, fishstat, n=10, year=2023
):
    top_species_dfs = {}

    for ag in ags:
        areas = get_numbers_from_string(ag)

        fs_area_mask = fishstat["Area"].isin(areas)
        sa_area_mask = stock_assessments["Analysis Group"] == ag

        species_in_area = stock_assessments[sa_area_mask][
            "ASFIS Scientific Name"
        ].unique()
        species_mask = fishstat["ASFIS Scientific Name"].isin(species_in_area) & (
            fishstat["ASFIS Scientific Name"] != "Actinopterygii"
        )

        top_species = (
            fishstat[fs_area_mask & species_mask]
            .groupby("ASFIS Scientific Name")[year]
            .sum()
            .nlargest(n)
            .reset_index()
        )
        top_species[2023] /= 1e3
        top_species = top_species.rename(
            columns={2023: "Landings 2023 (in thousand tonnes, live weight equivalent)"}
        )

        top_species_list = list(top_species["ASFIS Scientific Name"])

        sa_top_species_mask = stock_assessments["ASFIS Scientific Name"].isin(
            top_species_list
        )

        sbn = compute_status_by_number(
            stock_assessments[sa_top_species_mask & sa_area_mask],
            "ASFIS Scientific Name",
        )
        sbn.loc[sbn["ASFIS Scientific Name"] == "Global", "ASFIS Scientific Name"] = (
            "Total"
        )

        sl_area_mask = stock_landings["Analysis Group"] == ag
        sl_top_species_mask = stock_landings["ASFIS Scientific Name"].isin(
            top_species_list
        )

        sbl = compute_weighted_percentages(
            stock_landings[sl_area_mask & sl_top_species_mask],
            key="ASFIS Scientific Name",
        )
        sbl.columns = [col[1].replace("(Mt)", "(Kt)") for col in sbl.columns]
        sbl.loc[:, [col for col in sbl.columns if "(Kt)" in col]] *= 1e3
        sbl = sbl.reset_index()

        comb = pd.merge(top_species, sbn, on="ASFIS Scientific Name")
        comb = pd.merge(
            comb,
            sbl,
            on="ASFIS Scientific Name",
            suffixes=(" by Number", " by Landings"),
        )

        top_species_dfs[ag] = comb

    return top_species_dfs
